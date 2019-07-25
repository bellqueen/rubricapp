from flask import Flask, render_template, redirect, url_for, request, send_file
from flask_bootstrap import Bootstrap
from flask_wtf import FlaskForm 
from wtforms import StringField, PasswordField, BooleanField
from flask_wtf.file import FileField, FileAllowed, FileRequired
from wtforms.validators import InputRequired, Email, Length
from flask_sqlalchemy  import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import smtplib
from email.message import EmailMessage
from os.path import basename
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import os
#shutil used to delete whole directory(folder)
import shutil
import openpyxl
from openpyxl import load_workbook
from _datetime import datetime
import json

app = Flask(__name__)
app.config['SECRET_KEY'] = 'Thisissupposedtobesecret!'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///C:/Users/bell7/account.db'
bootstrap = Bootstrap(app)
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'


# Settings of Directory ======================================================================================================

#SET THE BASE DIRECTORY
os.chdir('D:/EDUsample/Accounts')
base_directory = os.getcwd()

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(15), unique=True)
    email = db.Column(db.String(50), unique=True)
    instructor = db.Column(db.String(2))
    password = db.Column(db.String(80))

class Permission(UserMixin, db.Model):
    project_id = db.Column(db.String(50), primary_key=True)
    owner = db.Column(db.String(50), primary_key=True)
    shareTo = db.Column(db.String(50), primary_key=True)
    project = db.Column(db.String(30), primary_key=True)
    status = db.Column(db.String(30))

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

class LoginForm(FlaskForm):
    username = StringField('username', validators=[InputRequired(), Length(min=4, max=15)])
    password = PasswordField('password', validators=[InputRequired(), Length(min=8, max=80)])
    remember = BooleanField('remember me')

class RegisterForm(FlaskForm):
    email = StringField('email', validators=[InputRequired(), Email(message='Invalid email'), Length(max=50)])
    username = StringField('username', validators=[InputRequired(), Length(min=4, max=15)])
    password = PasswordField('password', validators=[InputRequired(), Length(min=8, max=80)])
    instructor = BooleanField('instructor')

class ProjectForm(FlaskForm):
    project_name = StringField('project name', validators=[InputRequired(), Length(min=3, max=10)])
    project_description = StringField('description', validators=[InputRequired(), Length(min=0, max=50)])
    # group_file = FileField('group file',validators = [FileRequired(),FileAllowed(JSON, 'Json only')]
    group_file = FileField('group file')
    student_file = FileField('students file')
    grading_criteria = FileField('grading criteria')

# class EmailForm(FlaskForm):
#     email = StringField('email', validators=[InputRequired(), Email(message='Invalid email'), Length(max=50)])
#     password = PasswordField('password', validators=[InputRequired(), Length(min=1, max=80)])

# class EvaluationForm(FlaskForm):
#     evaluation_name = StringField('evaluation name', validators=[InputRequired(), Length(min=3, max=10)])

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()

    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user:
            if check_password_hash(user.password, form.password.data):
                login_user(user, remember=form.remember.data)
                # instructor jump to instructor page, student jump to student page
                if(user.instructor == "1"):
                    return redirect(url_for('instructor_dashboard'))
                else:
                    return redirect(url_for('student_dashboard'))
            else:
                return render_template('login.html', msg="password not correct", form=form)
        else:
            return render_template('login.html', msg="user doesn't exist", form=form)
        #return '<h1>' + form.username.data + ' ' + form.password.data + '</h1>'

    return render_template('login.html', msg="", form=form)

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    form = RegisterForm()
    if form.validate_on_submit():
        # check if the user and email has existed in the database
        check_username = User.query.filter_by(username=form.username.data).first()
        check_email = User.query.filter_by(email=form.email.data).first()
        if check_username:
            return render_template('signup.html', form=form, msg="Warning !!! The username has existed")
        elif check_email:
            return render_template('signup.html', form=form, msg="Warning !!! The email has been used")
        else:
            hashed_password = generate_password_hash(form.password.data, method='sha256')
            new_user = User(username=form.username.data, email=form.email.data, password=hashed_password, instructor=form.instructor.data)
            db.session.add(new_user)
            db.session.commit()

            #After making sure that the new user is created, the user's private folder can be created by using the user name

            path_to_user_folder = "{}/{}".format(base_directory, new_user.username)
            os.mkdir(path_to_user_folder)

            return redirect(url_for('login'))
            #return '<h1>' + form.username.data + ' ' + form.email.data + ' ' + form.password.data + '</h1>'

    return render_template('signup.html', form=form, msg="")

@app.route('/instructor_dashboard')
@login_required
def instructor_dashboard():
    #Load all projects to instructor_dashboard
    #Find all projects in User's private folder by using current user

    path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    project_list = []
    for project in os.listdir(path_to_current_user):
        project_list.append(project)
    project_len = len(project_list)

    return render_template('instructor_dashboard.html', name=current_user.username, project_list=project_list, project_len = project_len)

@app.route('/project_profile', methods=["POST", "GET"])
@login_required
def project_profile():
    #prepare list of project and map of project like {project1: [eva1, eva2, eva3]}
    #project1:{eva1: [Ricky, Jason]}
    #project1: [Ricky, Json] (share to whom)
    path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    project_set_map = {}
    project_dic_map = {}
    project_permission_map = {}
    list_of_all_projects = Permission.query.filter_by(owner=current_user.username).all()
    project_list = Permission.query.filter_by(owner=current_user.username, shareTo=current_user.username).all()
    list_of_shareTo_permission = [item for item in list_of_all_projects if item not in project_list]
    for project in list_of_shareTo_permission:
        if project.project in list(project_permission_map.keys()):
            project_permission_map[project.project].append(project)
        else:
            project_permission_map[project.project] = [project]

    for project in project_list:
        path_to_evaluation_file = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username, project.project)
        evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_file)
        evaluation_worksheet = evaluation_workbook['eva']
        list_of_eva = select_by_col_name('eva_name', evaluation_worksheet)
        set_of_eva = set(list_of_eva)
        project_set_map[project.project] = set_of_eva
        # get all owner
        dic_of_eva = {}
        for eva in set_of_eva:
            owners = set()
            temp_eva = select_row_by_group_id("eva_name", eva, evaluation_worksheet)
            for row in temp_eva:
                owners.add(row['owner'])
            dic_of_eva[eva] = owners
        project_dic_map[project.project] = dic_of_eva

    return render_template("project_profile.html", project_list=project_list, shareTo_permission=list_of_shareTo_permission, project_set_map=project_set_map, project_dic_map=project_dic_map)

@app.route('/delete_eva/<string:project_id>/<string:evaluation>/<string:owner>', methods=['GET', 'POST'])
@login_required
def delete_eva(project_id, evaluation, owner):
    project = Permission.query.filter_by(project_id=project_id)
    print(project_id)
    path_to_evaluation_xlsx = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username, project.project)
    evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_xlsx)
    evaluation_worksheet = evaluation_workbook['eva']
    group_worksheet = evaluation_workbook['group']

    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)
    for group in group_col:
        delete_index = select_index_by_group_eva_owner(evaluation, group, owner, evaluation_worksheet)
        print(delete_index)
        evaluation_worksheet.delete_rows(delete_index, 1)
    evaluation_workbook.save(path_to_evaluation_xlsx)
    return redirect(url_for("/project_profile"))

@app.route('/delete_project/<string:project_id>', methods=['GET', 'POST'])
@login_required
def delete_project(project_id):
    project = Permission.query.filter_by(project_id=project_id).first()
    permission_to_delete = Permission.query.filter_by(project=project.project).all()
    path_to_current_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
    if os.path.exists(path_to_current_project):
        shutil.rmtree(path_to_current_project)
        #after delete the folder, delete all the permissions that were send from the project
        for permission in permission_to_delete:
            query = Permission.delete().where(Permission.project_id == permission.project_id)
            query.execute()
        msg = "project has been deleted successfully"
    else:
        msg = "something went wrong with the project"

    return redirect(url_for("/project_profile"))

@app.route('/update_permission/<string:project_id>/<string:authority>', methods=["GET", "POST"])
@login_required
def update_permission(project_id, authority):
    try:
        query = Permission.update().where(Permission.project_id == project_id).values(status=authority)
        query.execute()
        msg = "successfully updated authority"
    except Exception as e:

        msg = "failure to update authority, {}".format(e)

    return redirect(url_for("/project_profile"))

@app.route('/create_permission/<string:project_id>', methods=["GET", "POST"])
@login_required
def create_permission(project_id):
    username = request.form.get('username', " ")
    authority = request.form['authority']
    pending_authority = "pending|{}".format(authority)
    try:
        project = Permission.query.filter_by(project_id=project_id)
        permission_projectid = "{}{}{}{}".format(current_user.username, username, project.project, authority)
        query = Permission.insert().values(project_id=permission_projectid, owner=current_user.username, shareTo=username, status=pending_authority)
        query.execute()
        msg = "successfully created authority"
    except Exception as e:

        msg = "failure to create authority, {}".format(e)

    return redirect(url_for("/project_profile"))


@app.route('/modify_group/<string:project>')
@login_required
def modify_group(project):
    return


@app.route('/modify_group_receiver/<string:project>')
@login_required
def modify_group_receiver(project):
    return

@app.route('/instructor_project', methods=["POST","GET"])
@login_required
def instructor_project():
    #Load All project to instructor_project
    #Add a 'create new project' button
    # path_to_current_user = "{}/{}".format(base_directory, current_user.username)
    # project_list = []
    # for project in os.listdir(path_to_current_user):
    #     project_list.append(project)
    # project_len = len(project_list)

    #Load All project and shared project from database
    list_of_all_projects = Permission.query.filter_by(shareTo=current_user.username).all()
    list_of_personal_projects = Permission.query.filter_by(owner=current_user.username, shareTo=current_user.username).all()
    # list_of_shared_project = [item for item in list_of_all_projects if item not in list_of_personal_projects]
    list_of_shared_project = []
    for project in list_of_all_projects:
        for personal_project in list_of_personal_projects:
            if project.project_id != personal_project.project_id:
                list_of_shared_project.append(project)
    personal_project_len = len(list_of_personal_projects)
    shared_project_len = len(list_of_shared_project)


    return render_template('instructor_project.html', personal_project_list = list_of_personal_projects, shared_project_list = list_of_shared_project, personal_project_len=personal_project_len, shared_project_len=shared_project_len)

@app.route('/create_project', methods=["POST","GET"])
@login_required
def create_project():
    #Request from file by WTF
    #Create a new project folder under 'path_to_current_user'
    #save files in new folder and build a evaluation doc depending on json file
    form = ProjectForm()

    if form.validate_on_submit():
        #create project folder
        path_to_current_user_project = "{}/{}/{}".format(base_directory, current_user.username, form.project_name.data)
        os.mkdir(path_to_current_user_project)
        #save group file and grading criteria
        path_to_group_file_stored = "{}/".format(path_to_current_user_project)
        group_file_filename = "group.xlsx"
        form.group_file.data.save(path_to_group_file_stored + group_file_filename)
        path_to_student_file_stored = "{}/".format(path_to_current_user_project)
        student_file_filename = "student.xlsx"
        form.student_file.data.save(path_to_student_file_stored + student_file_filename)
        grading_criteria_filename = "json.json"
        form.grading_criteria.data.save(path_to_group_file_stored + grading_criteria_filename)
        #creating evaluation doc based on grading criteria json file

        #copy group sheet to evaluation doc
        path_to_group_file = "{}/group.xlsx".format(path_to_current_user_project)
        group_file_workbook = load_workbook(path_to_group_file)
        group_file_worksheet = group_file_workbook['Sheet1']
        path_to_student_file = "{}/student.xlsx".format(path_to_current_user_project)
        student_file_workbook = load_workbook(path_to_student_file)
        student_file_worksheet = student_file_workbook['Sheet1']

        path_to_evaluation = "{}/evaluation.xlsx".format(path_to_current_user_project)
        evaluation_workbook = openpyxl.Workbook()
        evaluation_group = evaluation_workbook.create_sheet('group')
        evaluation_description = evaluation_workbook.create_sheet('description')
        evaluation_student = evaluation_workbook.create_sheet('students')
        project_description = [form.project_name.data, form.project_description.data]
        evaluation_description.append(project_description)
        copy_all_worksheet(evaluation_group, group_file_worksheet)
        copy_all_worksheet(evaluation_student, student_file_worksheet)
        #create EVA depending on the json file
        evaluation_eva = evaluation_workbook.create_sheet('eva')
        #open json file and load json
        path_to_json_file="{}/json.json".format(path_to_current_user_project)
        with open(path_to_json_file, 'r')as f:
            json_data = json.loads(f.read(), strict=False)
        # The group id, eva_name, date are defults
        tags_to_append = ['group_id', 'eva_name', 'owner', 'date', 'students']
        for category in json_data['category']:
            category_name = (category['name'])
            for section in category['section']:
                #instructors don't care about the text value, the text values will only be send to students.
                if section['type'] != 'text':
                    value_to_append = "{}|{}".format(category_name, section['name'])
                    tags_to_append.append(value_to_append)
        tags_to_append.append("comment")
        tags_to_append.append("last_updates")
        evaluation_eva.append(tags_to_append)

        evaluation_workbook.save(path_to_evaluation)

        #create permission to owener himself
        project_id = "{}{}{}{}".format(current_user.username, current_user.username, form.project_name.data, 'full')
        self_permission = Permission(project_id=project_id, owner=current_user.username, shareTo=current_user.username, project=form.project_name.data, status='full')
        db.session.add(self_permission)
        db.session.commit()

        return redirect(url_for("instructor_project"))



    return render_template('create_project.html', form = form)


def copy_all_worksheet(copy_to, copy_from):
    for row in range(0, len(list(copy_from.iter_rows()))):
        for col in range(0, len(list(copy_from.iter_cols()))):
            copy_to.cell(row=row+1, column=col+1).value = copy_from.cell(row=row+1, column=col+1).value

@app.route('/load_project/<string:project_id>/<string:msg>',methods=["GET"])
@login_required
def load_project(project_id, msg):
    #get project by project_id
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_evaluation_xlsx = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username, project.project)
    evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_xlsx)
    evaluation_worksheet = evaluation_workbook['eva']
    list_of_eva = select_by_col_name('eva_name', evaluation_worksheet)
    set_of_eva = set(list_of_eva)

    # get all owner
    dic_of_eva = {}
    for eva in set_of_eva:
        owners = set()
        temp_eva = select_row_by_group_id("eva_name", eva, evaluation_worksheet)
        for row in temp_eva:
            owners.add(row['owner'])
        dic_of_eva[eva] = owners
    return render_template("project.html", project=project, data_of_eva_set=set_of_eva, data_of_eva_dic=dic_of_eva, msg=msg, useremail=current_user.email)

# @app.route('/create_new_evaluation/<string:project_name>')
# @login_required
# def create_new_evaluation(project_name):
#     return render_template('create_new_evaluation.html', project_name=project_name)

@app.route('/create_evaluation/<string:project_id>', methods=['GET', 'POST'])
@login_required
def create_evaluation(project_id):
    # get project by id
    project = Permission.query.filter_by(project_id=project_id).first()
    #load group columns and evaluation worksheet
    evaluation_name = request.form['evaluation_name']
    evaluation_desc = request.form.get('evaluation_description', " ")
    path_to_load_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    eva_workbook = load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']
    eva_worksheet = eva_workbook['eva']
    students_worksheet = eva_workbook['students']
    desc_worksheet = eva_workbook['description']
    desc_worksheet.append([evaluation_name, evaluation_desc])


    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)

    #get all students by students
    students = get_students_by_group(group_worksheet, students_worksheet)
    #create a empty row for each group in the new evaluation
    for group in group_col:
        students_name = []
        #couple is [email, student_name]
        for student_couple in students[str(group)]:
            students_name.append(student_couple[1])
        row_to_insert = new_row_generator(str(group), students_name, evaluation_name, eva_worksheet)
        print(row_to_insert)
        eva_worksheet.append(row_to_insert)
    eva_workbook.save(path_to_evaluation_file)
    msg = "New Evaluation has been created successfully"

    return redirect(url_for('jump_to_evaluation_page', project_id=project.project_id, evaluation_name=evaluation_name, owner=current_user.username, msg=msg))

@app.route('/jump_tool/<string:project_id>/<string:evaluation_name>', methods=["GET", "POST"])
@login_required
def jump_tool(project_id, evaluation_name):
    owner = request.form['owner']
    msg = "Connect to Evaluation Successfully"
    return redirect(url_for('jump_to_evaluation_page', project_id=project_id, evaluation_name=evaluation_name, owner=owner, msg=msg))

@app.route('/jump_to_evaluation_page/<string:project_id>/<string:evaluation_name>/<string:owner>/<string:msg>', methods=["GET", "POST"])
@login_required
def jump_to_evaluation_page(project_id, evaluation_name, owner, msg):
    #get project by project_id
    project = Permission.query.filter_by(project_id=project_id).first()
    #prepare the json data and group numbers before it jumps to evaluation page
    path_to_load_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
    with open ("{}/json.json".format(path_to_load_project), 'r')as f:
        json_data = json.loads(f.read(), strict=False)
    eva_workbook = load_workbook("{}/evaluation.xlsx".format(path_to_load_project))
    group_worksheet = eva_workbook['group']
    students_worksheet = eva_workbook['students']

    #data of groups
    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)

    #check if evaluation exists in the worksheet
    eva_worksheet = eva_workbook['eva']

    #Transform ROWS in worksheet to DICTIONARY
    new_row = {}
    first_row = list(eva_worksheet.iter_rows())[0]
    for tag in first_row:
        new_row[tag.value] = ""

    temp_eva = select_row_by_group_id("eva_name", evaluation_name, eva_worksheet)

    eva_to_edit = {}

    for group in group_col:
        for row in temp_eva:
            if str(group) == str(row['group_id']) and str(row['owner'] == owner):
                eva_to_edit[str(group)] = row

    students = get_students_by_group(group_worksheet, students_worksheet)

    return render_template("evaluation_page.html",  project=project, json_data=json_data, group_col=group_col, msg=msg, evaluation_name=evaluation_name, edit_data=eva_to_edit, owner=owner, students=students)



# @app.route('/evaluation_page', methods=['GET', 'POST'])
# @login_required
# def evaluation_page():
#     return render_template("evaluation_page.html")



@app.route('/evaluation_receiver/<project_id>/<string:evaluation_name>/<string:group_div>/<string:owner>', methods=["GET", "POST"])
@login_required
def evaluation_page(project_id, evaluation_name, group_div, owner):
    #receive all the data and insert them into xlsx
    #group id, evaluation name, date time is constant
    row_to_insert = []
    group_id = group_div
    date = datetime.today().strftime('%d/%m/%Y')
    row_to_insert.append(group_id)
    row_to_insert.append(evaluation_name)
    row_to_insert.append(owner)
    row_to_insert.append(date)
    student_list = request.form['student']
    row_to_insert.append("|".join(student_list))

    #The rest are variables from TW
    # get project by project_id
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
    with open("{}/json.json".format(path_to_load_project), 'r')as f:
        json_data = json.loads(f.read(), strict=False)
    for category in json_data['category']:
        category_name = category['name']
        ratings_name = '{}|Ratings'.format(category_name)
        oc_name = '{}|Observable Characteristics'.format(category_name)
        # sg now is text
        # sg_name = '{}|Suggestions'.format(category_name)
        Ratings = request.form.get(ratings_name, " ")
        row_to_insert.append(Ratings)
        Observable_Characteristics = request.form.getlist(oc_name)
        if len(Observable_Characteristics) != 0:
            Observable_Characteristics_str = ';'.join(Observable_Characteristics)
        else:
            Observable_Characteristics_str = " "
        row_to_insert.append(Observable_Characteristics_str)
        # Suggestions = request.form.getlist(sg_name)
        # if len(Suggestions) != 0:
        #     Suggestions_str = ';'.join(Suggestions)
        # else:
        #     Suggestions_str = " "
        # row_to_insert.append(Suggestions_str)
    comment = request.form.get('comment', " ")
    row_to_insert.append(comment)

    # After everything enrolls, insert the row to evaluation worksheet
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    evaluation_workbook = load_workbook(path_to_evaluation_file)
    evaluation_worksheet = evaluation_workbook['eva']
    #change the last update by append the current user
    index = int(select_index_by_group_eva_owner(evaluation_name, group_id, owner, evaluation_worksheet))
    last_update = select_by_col_name('last_updates', evaluation_worksheet)[index-2]
    last_update = "{}|{}".format(last_update, current_user.username)
    row_to_insert.append(last_update)

    #delete the old row by index
    evaluation_worksheet.delete_rows(index, 1)
    evaluation_worksheet.append(row_to_insert)
    evaluation_workbook.save(path_to_evaluation_file)
    msg = "The grade has been updated successfully"
    return redirect(url_for('jump_to_evaluation_page', project_id=project_id, evaluation_name=evaluation_name, owner=owner, msg=msg))

# @app.route('/instructor_grade/<string:project_name>', methods=['GET','POST'])
# @login_required
# def instuctor_grade(project_name):
#     path_to_evaluation_xlsx = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username, project_name)
#     evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_xlsx)
#     evaluation_worksheet = evaluation_workbook['eva']
#     list_of_eva = select_by_col_name('eva_name', evaluation_worksheet)
#     set_of_eva = set(list_of_eva)
#     return render_template('instructor_grade.html', project_name=project_name, data_of_eva_set=set_of_eva)

# @app.route('/instructor_evaluations/<string:project_name>', methods=['GET','POST'])
# @login_required
# def instructor_evaluations(project_name):
#     path_to_evaluation_xlsx = "{}/{}/{}/evaluation.xlsx".format(base_directory, current_user.username, project_name)
#     evaluation_workbook = openpyxl.load_workbook(path_to_evaluation_xlsx)
#     evaluation_worksheet = evaluation_workbook['eva']
#     list_of_eva = select_by_col_name('eva_name', evaluation_worksheet)
#     set_of_eva = set(list_of_eva)
#     return render_template("instructor_evaluations.html", project_name=project_name, data_of_eva_set=set_of_eva)

#url address for download_page
#download_page include only data for one group
@app.route('/download_page/<string:project_id>/<string:evaluation_name>/<string:group>/<string:owner>', methods=['GET', 'POST'])
@login_required
def download_page(project_id, evaluation_name, group, owner):
    # get project by project_id
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    eva_workbook = load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']
    eva_worksheet = eva_workbook['eva']
    students_worksheet = eva_workbook['students']

    with open("{}/json.json".format(path_to_load_project), 'r')as f:
        json_data = json.loads(f.read(), strict=False)
    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)
    index_of_group = int(select_index_by_group_eva_owner(evaluation_name, owner, group, eva_worksheet))
    grade_map = select_map_by_index(index_of_group, eva_worksheet)

    students_in_one_group = get_students_by_group(group_worksheet, students_worksheet)[group]
    msg = "Downloaded grade group{}".format(group)

    return render_template("download_page.html", project=project, json_data=json_data, group=group, msg=msg, evaluation_name=evaluation_name, edit_data=grade_map, owner=owner, students= students_in_one_group)

#download xlsx file or html file
@app.route('/download/<string:project_id>/<string:evaluation_name>', methods=['GET', 'POST'])
@login_required
def download(project_id, evaluation_name):
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    eva_workbook = load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']
    eva_worksheet = eva_workbook['eva']
    students_worksheet = eva_workbook['students']
    students = get_students_by_group(group_worksheet, students_worksheet)
    with open("{}/json.json".format(path_to_load_project), 'r')as f:
        json_data = json.loads(f.read(), strict=False)
    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)
    if evaluation_name == "all":
        try:
            filename = "{}_{}.xlsx".format(project.project, evaluation_name)
            return send_file(path_to_evaluation_file, attachment_filename=filename)
            # msg = "Successfully downloaded"
            # return redirect(url_for('load_project', project_name=project_name, msg=msg))
        except Exception as e:
            print(str(e))
    else:
        # response = urllib.request.urlopen(url_for('jump_to_evaluation_page', project_name=project_name, evaluation_name=evaluation_name))
        # html = response.read()
        # print(html)
        new_row = {}
        first_row = list(eva_worksheet.iter_rows())[0]
        for tag in first_row:
            new_row[tag.value] = ""
        # grab the data that used for this evaluation
        owner = request.form['owner']
        temp_eva = select_row_by_group_id("eva_name", evaluation_name, eva_worksheet)
        eva_to_edit = {}
        for group in group_col:
            for row in temp_eva:
                if str(group) == str(row['group_id']) and str(owner) == str(row['owner']):
                    eva_to_edit[str(group)] = row
        #store the data to a html file and send out(download)
        msg = ""
        path_to_html = "{}/{}_{}.html".format(path_to_load_project, project.project, evaluation_name)
        #remove the old file in case duplicated file existence
        if os.path.exists(path_to_html):
            os.remove(path_to_html)
        with open(path_to_html, 'w') as f:
            f.write(render_template("evaluation_page.html", project=project, json_data=json_data, group_col=group_col, msg=msg, evaluation_name=evaluation_name, edit_data=eva_to_edit, owner = owner, students=students))
        return send_file(path_to_html, as_attachment=True)

@app.route('/sendEmail/<string:project_id>/<string:evaluation_name>', methods=['GET', 'POST'])
@login_required
def sendEmail(project_id, evaluation_name):
    project = Permission.query.filter_by(project_id=project_id).first()
    path_to_load_project = "{}/{}/{}".format(base_directory, current_user.username, project.project)
    path_to_evaluation_file = "{}/evaluation.xlsx".format(path_to_load_project)
    eva_workbook = load_workbook(path_to_evaluation_file)
    group_worksheet = eva_workbook['group']
    eva_worksheet = eva_workbook['eva']
    students_worksheet = eva_workbook['students']
    owner = request.form['owner_submit']
    with open("{}/json.json".format(path_to_load_project), 'r')as f:
        json_data = json.loads(f.read(), strict=False)
    # data of groups
    group_col = []
    for col_item in list(group_worksheet.iter_cols())[0]:
        if col_item.value != "groupid":
            group_col.append(col_item.value)

    #request data from EmailForm
    useremail = request.form['useremail']
    userpassword = request.form['userpassword']

    #set up server and log in email
    try:
        server = smtplib.SMTP(host='smtp-mail.outlook.com', port=587)
        server.starttls()
        server.login(useremail, userpassword)
        #send out emails by group
        for group in group_col:
            students_email = select_students_by_group(group, group_worksheet)
            index_of_group = int(select_index_by_group_eva_owner(evaluation_name, group, owner, eva_worksheet))
            grade_map = select_map_by_index(index_of_group, eva_worksheet)
            students_in_one_group = get_students_by_group(group_worksheet, students_worksheet)[group]
            #load download_page.html and store it to 'part' which will be attached to message in mail
            path_to_html = "{}/{}_{}_{}.html".format(path_to_load_project, project.project, evaluation_name, group)
            #write the download page html and automatically stored in local project
            msg = "Downloaded grade group{}".format(group)
            with open(path_to_html, 'w') as f:
                f.write(render_template("download_page.html", project=project, json_data=json_data,
                                        group=group, msg=msg, evaluation_name=evaluation_name,
                                        edit_data=grade_map, owner=owner, students=students_in_one_group))
            #load the download page to message

            with open(path_to_html, 'rb')as f:
                file_data = f.read()
                file_name = "{}_{}_{}.html".format(project.project, evaluation_name, group)
            for email in students_email:
                # create an instance of message
                if email is not None:
                    message = EmailMessage()
                    message['From'] = useremail
                    message['To'] = email
                    message['Subject'] = "{}|{}|{}".format(project.project, evaluation_name, group)
                    text = "send from {}, {} for group{}".format(project.project, evaluation_name, group)
                    message.set_content(text)
                    # message.attach(part)
                    message.add_attachment(file_data, maintype='html', subtype='html', filename=file_name)

                    server.send_message(message)
            #remove the html file after sending email
            #in case of duplicated file existence
            if os.path.exists(path_to_html):
                os.remove(path_to_html)
        server.close()
        msg = "Emails send out Successfully"
    except Exception as e:
        print('Something went wrong' + str(e))
        msg = "Something went wrong"

    return redirect(url_for('load_project', project_id=project_id, msg=msg))

@app.route('/student_dashboard')
@login_required
def student_dashboard():
    return render_template('student_dashboard.html', name=current_user.username)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

# def sendmail(from_email, password, to_email, subject, message):
#     msg = MIMEMultipart()
#     msg['From'] = from_email
#     msg['To'] = to_email
#     msg['Subject'] = subject
#     msg.attach(MIMEText(message, 'plain'))
#     try:
#         server = smtplib.SMTP(host='smtp-mail.outlook.com', port=587)
#         server.starttls()
#         server.login(from_email, password)
#         server.send_message(msg)
#         server.quit()
#         return "successfully"
#     except Exception as e:
#         print('Something went wrong' + str(e))
#         return "Something went wrong"

#select all elements by the column name
def select_by_col_name(col_name, worksheet):
    first_row = list(worksheet.iter_rows())[0]
    index_of_col = -1
    for variables in first_row:
        if variables.value == col_name:
            index_of_col = first_row.index(variables)
    if index_of_col == -1:
        return "col_name unmatched"
    else:
        col = list(worksheet.iter_cols())[index_of_col]
        data_by_col = []
        for data in range(1, len(col)):
            data_by_col.append(col[data].value)
        return data_by_col

#index is the row index in xsl
def select_map_by_index(index, worksheet):
    list_of_rows = list(worksheet.iter_rows())
    # row_to_return = []
    # for item in list_of_rows[index-1]:
    #     row_to_return.append(item.value)
    row_selected = list_of_rows[index-1]
    tags = list_of_rows[0]
    map_to_return = {}
    for tag_index in range(0, len(tags)):
        tag_value = tags[tag_index].value
        map_to_return[tag_value] = row_selected[tag_index].value
    return map_to_return

#index is the row index in xsl
def select_row_by_index(index, worksheet):
    list_of_rows = list(worksheet.iter_rows())
    row_to_return = []
    for item in list_of_rows[index-1]:
        row_to_return.append(item.value)
    return row_to_return
    
#group_id and eva_name and owner are primary key in eva worksheet
def select_index_by_group_eva_owner(eva_name, group_id, owner, worksheet):
    evaluation_list = select_by_col_name("eva_name", worksheet)
    group_list = select_by_col_name("group_id", worksheet)
    owner_list = select_by_col_name("owner", worksheet)
    for index in range(0, len(evaluation_list)):
        if evaluation_list[index] == eva_name and group_list[index] == group_id and owner_list[index] == owner:
            return index + 2
    return "nothing find"

def select_row_by_group_id(col_name, col_value, worksheet):
    #we suppose that the return rows are multiple
    rows_selected = []
    rows_in_worksheet = list(worksheet.iter_rows())
    # index in data_by_col == index in row
    data_by_col = select_by_col_name(col_name, worksheet)
    for index_data in range(0, len(data_by_col)):
        if data_by_col[index_data] == col_value:
            map_to_append = {}
            #tranfer the item to item.value
            for item in rows_in_worksheet[index_data+1]:
                #create a map contains keys-values : like "group_id" = 1
                index_of_item = rows_in_worksheet[index_data+1].index(item)
                tag_of_item = rows_in_worksheet[0][index_of_item].value
                map_to_append[tag_of_item] = item.value
            #record each row by its group_id
            #for example : 1:{group_id = 1, date= x/y/z, eva_name = eva3, .....}.
            rows_selected.append(map_to_append)
    #if nothing matched, return the empty array;
    return rows_selected

#select all students in the group
#implement in group worksheet
def select_students_by_group(group, worksheet):
    students = []
    groups = select_by_col_name("groupid", worksheet)
    index_of_group = groups.index(group)+2
    row_by_index = list(worksheet.iter_rows())[index_of_group-1]
    #the first element in row_by_index is groupid
    for student in row_by_index[1:]:
        students.append(student.value)
    return students
        
#generate an empty map with group name and evaluation name
def new_map_generator(group, eva_name, worksheet):
    map_to_return = {}
    tags_of_item = list(worksheet.iter_rows())[0]
    for tag in tags_of_item:
        if tag.value == 'group_id':
            map_to_return[tag.value] = group
        elif tag.value == 'eva_name':
            map_to_return[tag.value] = eva_name
        else:
            map_to_return[tag.value] = ''

    return map_to_return
def new_row_generator(group, students, eva_name, worksheet):
    row_to_return =[]
    tags_of_item = list(worksheet.iter_rows())[0]
    for tag in tags_of_item:
        if tag.value == 'group_id':
            row_to_return.append(group)
        elif tag.value == 'eva_name':
            row_to_return.append(eva_name)
        elif tag.value == 'date':
            date = datetime.today().strftime('%d/%m/%Y')
            row_to_return.append(date)
        elif tag.value == 'owner':
            row_to_return.append(current_user.username)
        elif tag.value == 'students':
            students_string = ",".join(students)
            row_to_return.append(students_string)
        elif tag.value == 'last_updates':
            row_to_return.append(current_user.username)
        else:
            row_to_return.append(" ")
    return row_to_return

def get_students_by_group(group_worksheet, students_worksheet):
    return_map = {}
    for group in list(group_worksheet.iter_rows())[1:]:
        return_map[group[0].value]=[]
        for email in group[1:]:
            if email.value is not None:
                index = select_by_col_name('Email', students_worksheet).index(email.value)
                student_name = select_by_col_name('Student',students_worksheet)[index]
                student_couple = [email.value, student_name]
                return_map[group[0].value].append(student_couple)
    return return_map

#After login===============================================================================================================================

if __name__ == '__main__':
    app.run(debug=True)